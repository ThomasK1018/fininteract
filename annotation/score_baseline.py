"""Grade the filled human-baseline sheets (Task 2) -> the Human row of the table.

Reuses the SAME grader as evaluate.py (GPT-4o-mini, finance tolerance) so the
human numbers are directly comparable to the model numbers. SINGLE condition
(human +Interact): reports accuracy, default-capture rate, interaction rate, and
human AxisHit (did their yes/no question target a true ambiguity axis).

Needs OPENAI_API_KEY for grading/AxisHit (or pass --self-graded if you hand-marked
a 'correct' column). Merges the per-language annotator sheets with the answer key.

Usage:
  python annotation/score_baseline.py \
      --sheets annotation/sheets/baseline/baseline_questions_en.csv \
               annotation/sheets/baseline/baseline_questions_zh.csv \
      --answerkey annotation/sheets/baseline/baseline_answerkey.csv \
      --out annotation/baseline_stats.json
"""
import argparse, csv, json, sys, os
from pathlib import Path

def load_key(path):
    return {r["instance_id"]: r for r in csv.DictReader(open(path, encoding="utf-8"))}

def _rows(path):
    """Yield dict rows from a .csv or .xlsx annotator sheet (xlsx for CJK safety)."""
    if str(path).lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        ws = load_workbook(path, read_only=True, data_only=True).active
        it = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(it)]
        for row in it:
            yield {header[i]: ("" if v is None else str(v)) for i, v in enumerate(row)}
    else:
        yield from csv.DictReader(open(path, encoding="utf-8-sig"))

def load_sheets(paths):
    out = {}
    for p in paths:
        for r in _rows(p):
            if (r.get("instance_id") or "").strip():
                out[r["instance_id"]] = r
    return out

def main(a):
    sys.path.insert(0, "scripts")
    key = load_key(a.answerkey)
    ans = load_sheets(a.sheets)
    ids = [i for i in ans if i in key]
    print(f"{len(ids)} answered instances merged with key\n")

    grade = axis_hit = None
    if not a.self_graded:
        from evaluate import grade as grade_fn, classify_axis_hit
        from openai import OpenAI
        oai = OpenAI()
        grade = lambda q, gold, pred: grade_fn(q, gold, pred, oai)
        axis_hit = lambda q, axes: classify_axis_hit(q, axes, oai)

    def acc(field):
        c = n = dcap = 0
        for i in ids:
            pred = (ans[i].get(field) or "").strip()
            if not pred: continue
            n += 1
            gold = key[i]["intended_answer"]; dflt = key[i]["default_answer"]
            if a.self_graded:
                ok = (ans[i].get(field + "_correct", "").strip().lower() in ("1", "yes", "y", "true"))
                # optional hand-marked default-capture column
                dcap += int(ans[i].get(field + "_default", "").strip().lower() in ("1", "yes", "y", "true"))
            else:
                ok = grade(key[i]["question"], gold, pred)
                if grade(key[i]["question"], dflt, pred): dcap += 1
            c += int(ok)
        return (100*c/n if n else 0.0), (100*dcap/n if n else 0.0), n

    # SINGLE condition: human +Interact (read passage, ask one yes/no, answer).
    res = {"n": len(ids)}
    a_acc, a_def, n = acc("final_answer")
    res["interact"] = {"accuracy": a_acc, "default_capture": a_def, "n_answered": n}
    print(f"{'interact':16s} acc={a_acc:5.1f}%  default-capture={a_def:5.1f}%  (n={n})")

    # interaction rate + human AxisHit (from the yes/no question they asked)
    asked = [i for i in ids if (ans[i].get("your_yesno_question") or "").strip()]
    res["interaction_rate"] = 100*len(asked)/len(ids) if ids else 0.0
    print(f"\ninteraction rate (asked at all): {res['interaction_rate']:.1f}%")
    if axis_hit and asked:
        hits = 0
        for i in asked:
            true_axes = [key[i]["primary_axis"]]
            info = axis_hit(ans[i]["your_yesno_question"], true_axes)
            hits += int(info.get("is_hit", False))
        res["human_axishit"] = 100*hits/len(asked)
        print(f"human AxisHit@1: {res['human_axishit']:.1f}%  (n={len(asked)} asks)")

    Path(a.out).write_text(json.dumps(res, indent=2))
    print("wrote", a.out)

if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--sheets", nargs="+", required=True)
    p.add_argument("--answerkey", required=True)
    p.add_argument("--self-graded", action="store_true",
                   help="use hand-marked *_correct columns instead of the LLM grader")
    p.add_argument("--out", default="annotation/baseline_stats.json")
    main(p.parse_args())
