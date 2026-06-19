"""Export the human-baseline sheets (Task 2: performance ceiling).

The human task mirrors the MODELS' task: given the ambiguous question and a
retrieved passage containing BOTH candidate values (intended + default,
shuffled, neither marked correct -- exactly what the model's `search` returns),
do you recognize the ambiguity and resolve it? This measures DISAMBIGUATION, not
recall, and makes the Human row directly comparable to the +Search / +Interact
model columns.

Produces THREE files for a stratified subset (default 50, language-balanced):
  1. baseline_questions_<lang>.csv  -- what the ANNOTATOR sees: question +
     retrieved_passage (both values, unlabeled) + blank answer columns. The
     resolving context C and the intended/default LABELS are withheld -> no leak.
  2. baseline_answerkey.csv         -- what the EXPERIMENTER holds: intended/
     default answers, context C, interpretations, used to (a) answer the
     annotator's yes/no question from C, and (b) grade later.

Protocol (per instance, ONE condition -- the human +Interact ceiling):
  Annotator reads the passage, asks ONE yes/no question, the experimenter answers
  it truthfully from C (Yes/No/IDK), then the annotator commits ONE final answer.
  One commitment per item (a prior two-condition design caused copy-through).

Usage:
  python annotation/make_baseline_sheet.py --n 50 --seed 11 \
      --outdir annotation/sheets/baseline
"""
import argparse, json, csv, random
from pathlib import Path


def write_xlsx(path, header, rows):
    """Write an .xlsx sheet (wrapped passage column) so Excel renders CJK correctly."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    wb = Workbook(); ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    widths = {"question": 42, "retrieved_passage": 70, "your_yesno_question": 38, "final_answer": 22}
    for ci, name in enumerate(header, start=1):
        col = ws.cell(row=1, column=ci).column_letter
        ws.column_dimensions[col].width = widths.get(name, 16)
        if name == "retrieved_passage":
            for ri in range(2, ws.max_row + 1):
                ws.cell(row=ri, column=ci).alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def nonleaky_passage(inst, rng):
    """Both evidence spans, shuffled, presented as unlabeled retrieved excerpts."""
    spans = [s for s in (inst.get("intended_evidence_span", ""),
                         inst.get("default_evidence_span", "")) if s]
    if len(spans) < 2:
        return (inst.get("passage_text", "") or (spans[0] if spans else ""))[:1200]
    rng.shuffle(spans)
    return "Retrieved excerpts (the query is under-specified; more than one is consistent):\n" + \
           "\n".join(f"[Excerpt {i+1}] {s}" for i, s in enumerate(spans))

def stratified(rows, n, seed):
    rng = random.Random(seed)
    buckets = {}
    for r in rows:
        buckets.setdefault((r["language"], (r.get("axes") or ["?"])[0]), []).append(r)
    for b in buckets.values():
        rng.shuffle(b)
    out, keys = [], sorted(buckets, key=lambda k: -len(buckets[k]))
    while len(out) < n and any(buckets[k] for k in keys):
        for k in keys:
            if buckets[k] and len(out) < n:
                out.append(buckets[k].pop())
    return out

def main(a):
    rows = [json.loads(l) for l in open(a.instances)]
    sample = stratified(rows, min(a.n, len(rows)), a.seed)
    outdir = Path(a.outdir); outdir.mkdir(parents=True, exist_ok=True)

    # 1) annotator-facing sheets, split by language (question + non-leaky passage).
    # SINGLE condition: read passage -> ask ONE yes/no question -> write ONE final
    # answer. One commitment per item (the earlier two-condition design caused
    # copy-through fatigue). This measures the human +Interact ceiling directly.
    for lang in sorted(set(r["language"] for r in sample)):
        sub = [r for r in sample if r["language"] == lang]
        path = outdir / f"baseline_questions_{lang}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["instance_id", "company", "question", "retrieved_passage",
                        "your_yesno_question", "experimenter_reply", "final_answer"])
            sheet_rows = []
            for idx, r in enumerate(sub):
                rng = random.Random(10_000 + idx)
                row = [r["instance_id"], r.get("company", ""), r["question"],
                       nonleaky_passage(r, rng), "", "", ""]
                w.writerow(row); sheet_rows.append(row)
        # also emit .xlsx -- Excel mangles UTF-8 CSV for CJK; xlsx renders ZH cleanly
        write_xlsx(outdir / f"baseline_questions_{lang}.xlsx",
                   ["instance_id", "company", "question", "retrieved_passage",
                    "your_yesno_question", "experimenter_reply", "final_answer"],
                   sheet_rows)
        print(f"wrote {path}(+.xlsx)  ({len(sub)} {lang} instances) -- annotator-facing, "
              f"question + both-value passage, single condition, no answer key")

    # 2) experimenter answer key (HOLD BACK from annotators)
    keypath = outdir / "baseline_answerkey.csv"
    with open(keypath, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["instance_id", "language", "primary_axis", "question",
                    "context_C", "intended_answer", "default_answer",
                    "intended_interpretation", "default_interpretation"])
        for r in sample:
            w.writerow([r["instance_id"], r["language"], (r.get("axes") or ["?"])[0],
                        r["question"], r["context"], r["answer"],
                        r.get("default_answer", ""),
                        json.dumps(r.get("intended_interpretation", {}), ensure_ascii=False),
                        json.dumps(r.get("default_interpretation", {}), ensure_ascii=False)])
    print(f"wrote {keypath}  -- EXPERIMENTER ONLY (gold answers + C). Do not show annotators.")

if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--instances", default="data/final/fininteract_v1.jsonl")
    p.add_argument("--n", type=int, default=50)
    p.add_argument("--seed", type=int, default=11)
    p.add_argument("--outdir", default="annotation/sheets/baseline")
    main(p.parse_args())
